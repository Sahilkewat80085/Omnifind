from datetime import datetime, timezone
from pathlib import Path
import openpyxl

from core.document_pipeline.extractors.base import BaseExtractor
from core.document_pipeline.models import (
    FormatMetadata,
    NormalizedBlock,
    NormalizedDocument,
    NormalizedTable,
)


class XlsxExtractor(BaseExtractor):
    """Extractor for Microsoft Excel (.xlsx) workbooks extracting sheets, tables, and formula values."""

    def extract(self, file_path: str) -> NormalizedDocument:
        path = Path(file_path)
        stat = path.stat()
        created_at = datetime.fromtimestamp(stat.st_ctime, tz=timezone.utc)
        modified_at = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)

        # Load with data_only=True to extract computed formula values
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        sheet_names = wb.sheetnames

        blocks: list[NormalizedBlock] = []
        tables: list[NormalizedTable] = []

        for sheet_idx, sheet_name in enumerate(sheet_names, start=1):
            sheet = wb[sheet_name]
            sheet_loc = f"Sheet: {sheet_name}"

            # Extract grid data
            raw_rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                # Filter out completely empty trailing rows
                str_cells = [str(cell).strip() if cell is not None else "" for cell in row]
                if any(str_cells):
                    raw_rows.append(str_cells)

            if not raw_rows:
                continue

            headers = raw_rows[0]
            data_rows = raw_rows[1:] if len(raw_rows) > 1 else []

            # Add structured table
            tables.append(
                NormalizedTable(
                    table_id=f"table-sheet-{sheet_idx}",
                    location=sheet_loc,
                    headers=headers,
                    rows=data_rows[:200],  # Bound table sample
                    caption=f"Spreadsheet Sheet: {sheet_name}",
                )
            )

            # Generate natural language summary block for tabular embedding
            col_summary = ", ".join([h for h in headers if h][:8])
            row_count = len(data_rows)
            sheet_summary_text = (
                f"Worksheet '{sheet_name}' containing {row_count} records. "
                f"Columns include: {col_summary}."
            )

            # Sample row previews for semantic context
            sample_previews = []
            for r in data_rows[:5]:
                paired = [f"{h}={val}" for h, val in zip(headers, r) if h and val]
                if paired:
                    sample_previews.append(", ".join(paired[:5]))

            if sample_previews:
                sheet_summary_text += " Sample rows: " + "; ".join(sample_previews) + "."

            blocks.append(
                NormalizedBlock(
                    text=sheet_summary_text,
                    location=sheet_loc,
                    heading_path=[sheet_name],
                    block_type="paragraph",
                )
            )

        wb.close()

        return NormalizedDocument(
            file_path=str(path.resolve()),
            file_name=path.name,
            file_type="xlsx",
            created_at=created_at,
            modified_at=modified_at,
            blocks=blocks,
            tables=tables,
            format_metadata=FormatMetadata(sheet_names=sheet_names),
            is_scanned=False,
            requires_ocr=False,
        )
